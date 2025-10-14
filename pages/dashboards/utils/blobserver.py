"""Utilities for fetching and parsing serology data from blobserver.

This module provides small helpers to fetch remote Excel files with retry
logic and parse the first worksheet into a list of dictionary records that can
be consumed by Django views and templates.
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional

import httpx
from openpyxl import load_workbook


DEFAULT_TIMEOUT = httpx.Timeout(10.0, connect=5.0)
DEFAULT_HEADERS = {"User-Agent": "pathogens-portal/requests"}

logger = logging.getLogger(__name__)


def get_with_retries(
    url: str,
    retries: int = 3,
    timeout: Optional[httpx.Timeout] = None,
    headers: Optional[Dict[str, str]] = None,
    user_agent: Optional[str] = None,
) -> httpx.Response:
    """GET a URL with retry and timeout policy.

    Args:
        url: Absolute URL to fetch.
        retries: Number of attempts before failing.
        timeout: Optional custom timeout. If not provided, a sensible default
            with separate connect timeout is used.
        headers: Optional HTTP headers to include in the request. These
            headers will override defaults if keys overlap.
        user_agent: Optional User-Agent string. If provided, it overrides the
            default; if ``headers`` also contains ``User-Agent``, that takes
            precedence over this argument.

    Returns:
        httpx.Response: Successful response object with a 2xx status.

    Raises:
        RuntimeError: If all retry attempts fail. The original exception from
            the last attempt is chained for debugging context.
    """
    timeout = timeout or DEFAULT_TIMEOUT
    # Build final headers with precedence: DEFAULT < user_agent arg < headers arg
    final_headers: Dict[str, str] = dict(DEFAULT_HEADERS)
    if user_agent:
        final_headers["User-Agent"] = user_agent
    if headers:
        final_headers.update(headers)
    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        logger.debug(
            "http_get_attempt",
            extra={
                "url": url,
                "attempt": attempt,
                "retries": retries,
                "user_agent": final_headers.get("User-Agent"),
            },
        )
        try:
            with httpx.Client(timeout=timeout, headers=final_headers) as client:
                response = client.get(url)
                response.raise_for_status()
                logger.debug(
                    "http_get_success",
                    extra={"url": url, "status_code": response.status_code, "content_length": len(response.content)},
                )
                return response
        except Exception as exc:
            last_error = exc
            logger.warning(
                "http_get_retry",
                extra={"url": url, "attempt": attempt, "retries": retries},
                exc_info=True,
            )
    logger.error("http_get_failed", extra={"url": url, "retries": retries, "error_type": type(last_error).__name__})
    raise RuntimeError(
        f"Failed to fetch URL after {retries} attempts: {url!r}"
    ) from last_error


def fetch_excel_first_sheet_as_records(
    url: str,
    retries: int = 3,
    headers: Optional[Dict[str, str]] = None,
    user_agent: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Parse the first worksheet of an Excel file into record dictionaries.

    The first row is treated as the header row. Blank header cells are ignored,
    and completely empty data rows are skipped. Values are normalised so that
    missing cells become empty strings.

    Args:
        url: Absolute URL of the Excel file to download.
        retries: Number of fetch attempts before failing.
        headers: Optional HTTP headers forwarded to the download request.
        user_agent: Optional User-Agent string to use for the request when
            ``headers`` does not already define one.

    Returns:
        List[Dict[str, Any]]: One dictionary per non-empty row, limited to the
        non-blank headers in the first row.
    """
    logger.debug("excel_fetch_start", extra={"url": url})
    response = get_with_retries(url, retries=retries, headers=headers, user_agent=user_agent)
    workbook = load_workbook(
        io.BytesIO(response.content), read_only=True, data_only=True
    )
    first_sheet_name = workbook.sheetnames[0]
    worksheet = workbook[first_sheet_name]

    row_iterator = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iterator)
    except StopIteration:
        logger.info("excel_no_rows", extra={"url": url, "sheet": first_sheet_name})
        return []

    headers: List[str] = []
    for header_cell in header_row:
        # Normalise headers to non-empty strings
        header_text = "" if header_cell is None else str(header_cell).strip()
        headers.append(header_text)

    normalised_headers = [h for h in headers if h]
    logger.debug(
        "excel_headers_parsed",
        extra={"url": url, "sheet": first_sheet_name, "header_count": len(normalised_headers)},
    )

    records: List[Dict[str, Any]] = []
    for data_row in row_iterator:
        # Build a dict limited to normalised headers
        record: Dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = data_row[index] if index < len(data_row) else None
            record[header] = "" if value is None else value
        # Skip fully empty rows
        if any(value not in ("", None) for value in record.values()):
            # Keep only normalised headers (drop any blanks)
            records.append({key: record.get(key, "") for key in normalised_headers})

    logger.info(
        "excel_rows_parsed",
        extra={"url": url, "sheet": first_sheet_name, "row_count": len(records)},
    )
    return records
